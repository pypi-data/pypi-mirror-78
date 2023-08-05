from TDhelper.generic.transformationType import transformation
from TDhelper.document.excel.FieldType import *
from types import FunctionType, MethodType, ModuleType
from openpyxl import load_workbook
import csv
import copy

class _AttributeOverride:
    def __init__(self, name, m_type):
        self._name = name
        self._type = m_type

    def __get__(self, instance, owen):
        return instance.__dict__[self._name]

    def __set__(self, instance, value):
        instance.__dict__[self._name] = transformation(value, self._type)

    def __delete__(self, instance):
        instance.__dict__.pop(self._name)

class Meta:
    file = None
    sheet = 'sheet1'
    extension = 'xlsx' 


class modelMeta(type):
    def __new__(cls, name, bases, dct):
        attrs = {'mapping': {}, 'Meta':Meta, '__exit__': __exit__, 'readLine': readLine,
                 'close': close, '__initExcelHandle__': __initExcelHandle__}
        for name, value in dct.items():
            if (not isinstance(dct[name], type)) and (not isinstance(dct[name], FunctionType)):
                if not name.startswith('__'):
                    if isinstance(dct[name], FieldType):
                        attrs['mapping'][name] = value.bindCol
                        attrs[name] = _AttributeOverride(name, value.fieldType)
                    else:
                        raise Exception('field type must is FieldType.')
            else:
                if isinstance(dct[name], type):
                    if name == 'Meta':
                        for attr_name in dct[name].__dict__:
                            if not attr_name.startswith('__'):
                                setattr(attrs['Meta'], attr_name, dct[name].__dict__[attr_name])
                else:
                    attrs[name] = value
        return super(modelMeta, cls).__new__(cls, name, bases, attrs)




def __initExcelHandle__(self):
    try:
        if self.Meta.file:
            m_extension= self.Meta.file.rsplit('.')[1]
            if m_extension== 'csv':
                self.Meta.extension= 'csv'
                self.__excelHandle__= open(self.Meta.file)
                self.__sheetHandle__= csv.reader(self.__excelHandle__)
                for v in self.__sheetHandle__:
                    self.__count__+=1
            elif m_extension== 'xlsx' or m_extension== 'xls':
                self.Meta.extension= 'xlsx'
                self.__excelHandle__ = load_workbook(self.Meta.file)
                self.__sheetHandle__ = self.__excelHandle__[self.Meta.sheet]
                self.__count__= self.__sheetHandle__.max_row
        else:
            raise Exception('meta file is None.')
    except Exception as e:
        raise e


def __exit__(self, exc_type, exc_value, exc_t):
    self.close()


def close(self):
    self.__excelHandle__ = None
    self.__sheetHandle__ = None

def getCount(self):
    if self.__sheetHandle__:
        return self.__count__
    else:
        raise Exception("Sheet is none.")

def rows(self):
    m_datas=[]
    m_offset=1
    while True:
        m_ret= self.readLine(m_offset)
        if m_ret:
            m_datas.append(m_ret)
        else:
            break
        m_offset+=1
    return m_datas

def readLine(self, lineOffset=1):
    if lineOffset < 1:
        lineOffset = 1
    if self.__sheetHandle__:
        if self.Meta.extension == 'xlsx':
            if lineOffset <= self.__sheetHandle__.max_row:
                rowdata = []
                column = self.__sheetHandle__.max_column+1
                for i in range(1, column):
                    cellvalue = self.__sheetHandle__.cell(
                        row=lineOffset, column=i).value
                    rowdata.append(cellvalue)
                for (name, value) in self.mapping.items():
                    if value <= len(rowdata):
                        setattr(self, name, rowdata[value-1])
                    else:
                        raise Exception('mapping error:(%s,%s)' % (name, value))
                self.__recordOffset__= lineOffset
                return self
            else:
                return None
        else:
            rowdata= []
            if lineOffset<=self.__recordOffset__:
                self.__recordOffset__= lineOffset-1
            for offset in range(self.__recordOffset__,lineOffset):
                try:
                    rowdata= next(self.__sheetHandle__)
                except Exception as e:
                    rowdata=[]
            self.__recordOffset__+=1
            if rowdata:
                for (name, value) in self.mapping.items():
                    if value <= len(rowdata):
                        attr_value=rowdata[value-1]
                        if attr_value=='None':
                            attr_value='0.00'
                        setattr(self, name, attr_value)
                    else:
                        raise Exception('mapping error:(%s,%s)' % (name, value))
                return self
            else:
                return None
    else:
        raise Exception('Sheet Handle is None.')
