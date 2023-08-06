# coding: utf-8


"""
    xlTables - Load/generate table data with Excel
    from python dictionary structures

    cunningr - 2020

    Requires openpyxl >= 2.6.2, jsonschema


"""

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from jsonschema import validate


def load_xl_db(db_file, flatten=False, squash=False, data_only=False, string_only=False, fill_empty=False):
    """
    Load Excel Database and convert tables to dictionary

    Args:
        db_file: (string) An Excel file containing data defined in named tables
        flatten: (bool) True or False (default).  Removes the Worksheet from the dictionary
            and appends to the Worksheet name to table name for dictionary keys.
        squash: (bool) True or False (default).  Assumes that all Worksheets contain only
            one table and therefore uses the worksheet name as the returned
            dictionary key. Only valid if flatten=True.
        data_only: (bool) True or False (default). Controlls if cell data value is loaded or formula
        string_only: (bool) True or False (default). Controlls if cell value is loaded as strings or
            "native" variable type (str, int, long, etc.)
        fill_empty: (bool) True or False (default). Controlls if empty cells are filled with ""

    Returns:
        The the full json object if found or 'None' is no object is found

    """
    wb = openpyxl.load_workbook(filename=db_file, data_only=data_only)

    sheets = wb.sheetnames
    workbook_dict = {}
    workbook_dict_flat = {}

    # Iterate over the sheets->tables in the workbook
    for sheet in sheets:
        sheet_name = wb[sheet].title
        workbook_dict.update({sheet_name: {}})
        for table in wb[sheet].tables.values():

            if flatten:
                if squash:
                    dictionary_name = '{}'.format(sheet_name)
                    table_dict = _build_dict_from_table(wb[sheet], table, name=dictionary_name, string_only=string_only, fill_empty=fill_empty)
                else:
                    dictionary_name = '{}'.format(table.name)
                    table_dict = _build_dict_from_table(wb[sheet], table, name=dictionary_name, string_only=string_only, fill_empty=fill_empty)
            else:
                table_dict = _build_dict_from_table(wb[sheet], table)

            workbook_dict_flat.update(table_dict)
            workbook_dict[sheet_name].update(table_dict)

    if flatten:
        return workbook_dict_flat
    else:
        return workbook_dict


def dump_db_to_xl(db_file_name, data, expand=False, compress=False, first_row_is_header=True, table_style='TableStyleMedium2', row_offset=2, col_offset=1):
    """
    Add a list of dictionaries (rows) as an Excel table.  By default the first row is the header.
    The dictionaries are list of key:value pairs constituting each row of the table.

    Args:
        db_file_name: (string) Name of Excel sheet to be created
        data: (dict) Dictionary of tables where each key represents a table with a list of table rows
        expand: (bool) Expands the tables one per worksheet
        compress: (bool) Compresses all tables on to a single worksheet
        first_row_is_header: (bool) True (default) or False. Currently ignored - first row must be 'stuffed'
        to contain all keys for column headers.
        table_style: (sting) An Excel table style
        row_offset: (int) Integer to determine number of spacer rows to add
        col_offset: (int) Integer to determine number of spacer columns to add

    Returns:
        Nothing

    """
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)

    if compress and not expand:
        _ws = wb.create_sheet('Tables')
    elif compress and expand:
        print('WARNING: Cannot compress and expand at the same time.  Compress flag will be ignored')

    for ws in data.keys():
        if not expand and not compress:
            _ws = wb.create_sheet(ws)

        for table in data[ws].keys():
            if expand:
                _ws = wb.create_sheet(table)

            header_keys = data[ws][table][0].keys()
            _header_row = {}
            for header in header_keys:
                _header_row.update({header: ''})

            data[ws][table].insert(0, _header_row)
            add_table_to_worksheet(
                _ws,
                table,
                data[ws][table],
                first_row_is_header=True,
                table_style=table_style,
                row_offset=row_offset, col_offset=col_offset
            )

    wb.save(db_file_name)


def add_table_to_worksheet(_work_sheet, name, data, first_row_is_header=True, table_style='TableStyleMedium2', row_offset=2, col_offset=1):
    """
    Add a list of dictionaries (rows) as an Excel table.  By default the first row is the header.
    The dictionaries are list of key:value pairs constituting each row of the table.

    Args:
        _work_sheet: (object) An openpyxl ws object (this must be changes to ws name only)
        name: (string) The name of the table.  Must be globally unique within the Excel workbook
        data: (list) Table data as a list of dictionaries with each element constituting a row of the table
        first_row_is_header: (bool) True (default) or False. Uses the first element of data as table headers
        table_style: (sting) An Excel table style
        row_offset: (int) Integer to determine number of spacer rows to add
        col_offset: (int) Integer to determine number of spacer columns to add

    Returns:
        Nothing

    """
    if first_row_is_header:
        column_headers = data.pop(0)
    else:
        print('first_row_is_header=False is currently not supported')
        exit()

    _start_table_data = _new_table_setup(_work_sheet, column_headers, row_offset=row_offset, col_offset=col_offset)
    new_table_end_col = _start_table_data[0]
    new_table_start_row = _start_table_data[1]

    last_data_row = _add_table_data(_work_sheet, column_headers, data, col_offset=col_offset)

    # Calculate new data refs and insert table
    _t_scol = _get_cell_column_letter(1 + col_offset)
    _t_ecol = _get_cell_column_letter(new_table_end_col)
    _scoord = '{}{}'.format(_t_scol, new_table_start_row)
    _ecoord = '{}{}'.format(_t_ecol, last_data_row)
    _t_ref = '{}:{}'.format(_scoord, _ecoord)
    tab = Table(displayName=name, ref=_t_ref)
    style = TableStyleInfo(name=table_style, showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    _work_sheet.add_table(tab)


def add_schema_table_to_worksheet(_work_sheet, name, schema, data=None, table_style='TableStyleMedium2', row_offset=2, col_offset=1):
    """
    Add a list of dictionaries (rows) as an Excel table using a schema.
    The schema should define the column headers and any data validation lists (enum).

    Rows of data (list) are passed separately (optional) and will be validated against the schema
    Rows are dictionaries with key:value pairs constituting each row of the table.
    Row keys much match column headers defined in the schema

    Args:
        _work_sheet: (object) An openpyxl ws object (this must be changes to ws name only)
        name: (string) The name of the table.  Must be globally unique within the Excel workbook
        data: (list) Table data as a list of dictionaries with each element constituting a row of the table
        schema: (dict) Dictionary representing a json schema (https://json-schema.org/).
        table_style: (string) An Excel table style
        row_offset: (int) Integer to determine number of spacer rows to add
        col_offset: (int) Integer to determine number of spacer columns to add

    Returns:
        Nothing

    """
    if 'description' in schema.keys():
        descr = schema['description']
    else:
        descr = None

    # Add new table headers at end of sheet
    # column_headers = list(schema['properties'].keys())
    column_headers = schema['properties']
    _start_table_data = _new_table_setup(_work_sheet, column_headers, descr=descr, row_offset=row_offset, col_offset=col_offset)
    new_table_end_col = _start_table_data[0]
    new_table_start_row = _start_table_data[1]

    # Build data validation based on column ID
    dv_dict = _add_column_data_validation(_work_sheet, column_headers, schema, col_offset=col_offset)

    # Add rows to sheet with data validation
    if data is not None:
        last_data_row = _add_table_data(_work_sheet, column_headers, data, schema=schema, dv_dict=dv_dict,
                                        col_offset=col_offset)
    else:
        data = _fill_row_data(schema)
        last_data_row = _add_table_data(_work_sheet, column_headers, data, schema=schema, dv_dict=dv_dict,
                                        col_offset=col_offset)

    # Calculate new data refs and insert table
    _t_scol = _get_cell_column_letter(1 + col_offset)
    _t_ecol = _get_cell_column_letter(new_table_end_col)
    _scoord = '{}{}'.format(_t_scol, new_table_start_row)
    _ecoord = '{}{}'.format(_t_ecol, last_data_row)
    _t_ref = '{}:{}'.format(_scoord, _ecoord)
    tab = Table(displayName=name, ref=_t_ref)
    style = TableStyleInfo(name=table_style, showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    _work_sheet.add_table(tab)

    return


def validate_data(schema, data):

    return _validate_data(schema, data)


def _new_table_setup(_work_sheet, headers, descr=None, row_offset=0, col_offset=0):

    if descr is not None:
        _start_col = 1 + col_offset
        _work_sheet.append({_start_col: descr})

        _, _end_row = _get_end_of_data(_work_sheet)
        # Insert row_offset above table description
        if row_offset is not 0:
            _work_sheet.insert_rows(_end_row, amount=row_offset)

        _end_col = len(headers) + col_offset
        _, _end_row = _get_end_of_data(_work_sheet)
        _work_sheet.merge_cells(start_row=_end_row, start_column=_start_col, end_row=_end_row, end_column=_end_col)
        _start_col_letter = _get_cell_column_letter(_start_col)
        descr_coord = '{}{}'.format(_start_col_letter, _end_row)
        _work_sheet[descr_coord].fill = PatternFill("solid", fgColor="ffff00")

    # Build the column headers with positional refs
    _add_headers = {}
    _add_descriptions = []
    for idx, header in enumerate(headers, start=1):
        _col = idx + col_offset
        _add_headers.update({_col: header})
        # Test is the headers are a schema dict and check for a description key
        if isinstance(headers[header], dict):
            if 'description' in headers[header].keys():
                _add_descriptions.append((header, _col, headers[header]['description']))

    # Add the headers
    _work_sheet.append(_add_headers)

    end_col, end_row = _get_end_of_data(_work_sheet)
    # Add any schema property descriptions as cell comments
    for _description in _add_descriptions:
        _col = _get_cell_column_letter(_description[1])
        _cell = '{}{}'.format(_col, end_row)
        _comment = Comment(_description[2], "xlTables")
        _work_sheet[_cell].comment = _comment

    if row_offset is not 0 and descr is None:
        _work_sheet.insert_rows(end_row, amount=row_offset)

    # Calculate end of worksheet, where the new table data will start
    abs_end_col, end_row = _get_end_of_data(_work_sheet)
    end_col = len(headers) + col_offset

    return end_col, end_row


def _get_end_of_data(_work_sheet):
    _current_dimensions = _work_sheet.calculate_dimension()
    _ref_start, _ref_end = _current_dimensions.split(':')
    new_table_start = _get_cell_coordinates(_ref_end)

    return new_table_start


def _add_column_data_validation(_work_sheet, headers, _schema, col_offset=0):
    dv_dict = {}
    for idx, column in enumerate(headers, start=1):
        if 'enum' in _schema['properties'][column].keys():
            _dv = _create_enum_dv(_schema['properties'][column]['enum'], allow_blank=True)
            _work_sheet.add_data_validation(_dv)
            _col = _get_cell_column_letter(idx + col_offset)
            dv_dict.update({_col: _dv})
        elif 'tref' in _schema['properties'][column].keys():
            _dv = _create_tref_dv(_schema['properties'][column]['tref'], allow_blank=True)
            _work_sheet.add_data_validation(_dv)
            _col = _get_cell_column_letter(idx + col_offset)
            dv_dict.update({_col: _dv})
        elif 'boolean' in _schema['properties'][column]['type']:
            _dv = _create_bool_dv(allow_blank=True)
            _work_sheet.add_data_validation(_dv)
            _col = _get_cell_column_letter(idx + col_offset)
            dv_dict.update({_col: _dv})

    return dv_dict


def _add_table_data(_work_sheet, headers, data, schema=None, dv_dict=None, col_offset=0):

    if schema is not None:
        _validate_data(schema, data)

    for row in data:
        _add_row = {}
        for idx, key in enumerate(headers, start=1):
            if key in row.keys():
                _col = idx + col_offset
                _add_row.update({_col: row[key]})
            elif "fillRow" in row.keys():
                _col = idx + col_offset
                _add_row.update({_col: ''})

        _work_sheet.append(_add_row)

        if dv_dict is not None:
            _current_dimensions = _work_sheet.calculate_dimension()
            _junk, _end = _current_dimensions.split(':')
            _new_row_idx = _get_cell_coordinates(_end)[1]
            for col, dv in dv_dict.items():
                cell = '{}{}'.format(col, _new_row_idx)
                dv.add(cell)

    _current_dimensions = _work_sheet.calculate_dimension()
    _junk, _end = _current_dimensions.split(':')
    _last_row_idx = _get_cell_coordinates(_end)[1]

    return _last_row_idx


def _fill_row_data(_schema):
    _fill_row = {}
    if 'properties' in _schema.keys():
        for key, value in _schema['properties'].items():
            if 'default' in value.keys():
                _fill_row.update({key: value['default']})

    if(len(_fill_row.keys())) == 0:
        _fill_row = {"fillRow": True}

    return [_fill_row]


def _validate_data(_schema, _data):

    for _row in _data:
        has_error = False
        # print(_row)
        # print(_schema)
        try:
            validate(instance=_row, schema=_schema)
        except Exception as e:
            has_error = True
            print('ROW: {}'.format(_row))
            error = 'ERROR:', e
            print(error)

    if has_error:
        print("Table data has schema errors")
        return False
    else:
        return True


def _create_enum_dv(values, allow_blank=True):

    # Remove 'null' values from JSON schema
    values = [x for x in values if x is not None]
    # Stringify values and create data-validation object
    _values = '"{}"'.format(','.join(values))
    dv = DataValidation(type="list", formula1=_values, allow_blank=allow_blank, errorStyle='warning')

    # Optionally set a custom error message
    dv.error = 'Entry not in the list'
    dv.errorTitle = 'Invalid Entry'

    return dv


def _create_tref_dv(_tref, allow_blank=True):

    # Stringify values and create data-validation object
    dv = DataValidation(type="list", formula1=_tref, allow_blank=allow_blank, errorStyle='warning')

    # Optionally set a custom error message
    dv.error = 'Entry not in the list'
    dv.errorTitle = 'Invalid Entry'

    return dv


def _create_bool_dv(allow_blank=True):

    # Stringify values and create data-validation object
    _values = '"{}"'.format(','.join(['TRUE', 'FALSE']))
    dv = DataValidation(type="list", formula1=_values, allow_blank=allow_blank)

    # Optionally set a custom error message
    dv.error = 'Entry not in the list'
    dv.errorTitle = 'Invalid Entry'

    return dv


def _get_cell_coordinates(cell):
    xy = openpyxl.utils.cell.coordinate_from_string(cell)
    col = openpyxl.utils.cell.column_index_from_string(xy[0])
    row = xy[1]

    return col, row


def _get_cell_column_letter(_col):
    _letter = openpyxl.utils.cell.get_column_letter(_col)

    return _letter


def _build_dict_from_table(_work_sheet, _table, name=None, string_only=False, fill_empty=False):
    """
    Internal function for building a dictionary from an Excel table definition

    Args:
        _work_sheet: Openpyxl worksheet object
        _table: Openpyxl table object (must be contained in _work_sheet object)
        name: The name to be used as the returned dictionary key (defaults to table.name)

    Returns:
        The the full json object if found or 'None' is no object is found

    """
    if name is None:
        name = _table.name

    # Get the cell range of the table
    _table_range = _table.ref

    _keys = []

    for _column in _table.tableColumns:
        _keys.append(_column.name)

    _num_columns = len(_keys)
    _row_width = len(_work_sheet[_table_range][0])
    if _num_columns != _row_width:
        print('ERROR: Key count and row elements are not equal' + _num_columns, _row_width)

    _new_dict = {name: {}}
    _rows_list = []

    for _row in _work_sheet[_table_range]:
        _row_dict = {}
        for _cell, _key in zip(_row, _keys):
            if _cell.value == _key:
                # Pass over headers where cell.value equal key
                pass
            else:
                if fill_empty == True and _cell.value == None:
                    _row_dict[_key] = ""
                elif string_only == True:
                    _row_dict[_key] = str(_cell.value).lstrip().rstrip()
                else:
                    _row_dict.update({_key:_cell.value})

        if bool(_row_dict):
            _rows_list.append(_row_dict)

    _new_dict[name] = _rows_list

    return _new_dict


if __name__ == "__main__":
    pass

