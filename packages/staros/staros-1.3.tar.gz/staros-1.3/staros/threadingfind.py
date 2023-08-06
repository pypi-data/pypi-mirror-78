import threading
import staros.connection
import staros.outputparsing
import re
import paramiko
import time
import openpyxl
import json

out = {}


def gethostname(host,port,username,password):
    data = staros.connection._getsshdata5(host, port, username, password, "show version\n")
    return staros.outputparsing._parse_hostname(data)


def daemon(host, port, username, password, msisdn):
    #print "start 1"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    #print data
    ppp = staros.outputparsing._getsubscribermsisdn(str(data))
    #print str(hoho) + " "+ str(ppp)
    if ppp.__len__() > 1:
        hoho = gethostname(host, port, username, password)
        out.update({hoho: ppp})

    #print "stop 1"


def daemon1imsi(host, port, username, password, imsi):
    #print "start 1"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers imsi " + imsi + "\n")
    #print data
    ppp = staros.outputparsing._getsubscribermsisdn(str(data))
    #print str(hoho) + " "+ str(ppp)
    if ppp.__len__()>1:
        hoho = gethostname(host, port, username, password)
        out.update({hoho : ppp})

    #print "stop 1"

def daemon1(host,port,username,password,msisdn):
    #print "start 2"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    ppp1 = staros.outputparsing._getsubscribermsisdn(str(data))
    hoho1 = gethostname(host,port,username,password)
    if ppp1['sessnum']!=0:
        out.update({hoho1: ppp1})
    #print "stop 2"


def daemon2(host,port,username,password,msisdn):
    #print "start 3"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    ppp2 = staros.outputparsing._getsubscribermsisdn(str(data))
    hoho2 = gethostname(host,port,username,password)
    if ppp2['sessnum']!=0:
        out.update({hoho2: ppp2})
    #print "stop 3"


def daemon3(host, port, username, password, msisdn):
    #print "start 4"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    ppp3 = staros.outputparsing._getsubscribermsisdn(str(data))
    hoho3 = gethostname(host, port, username, password)
    if ppp3['sessnum'] != 0:
        out.update({hoho3: ppp3})
    #print "stop 4"


def daemon4(host, port, username, password, msisdn):
    #print "start 5"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    ppp4 = staros.outputparsing._getsubscribermsisdn(str(data))
    hoho4 = gethostname(host,port,username,password)
    if ppp4['sessnum']!=0:
        out.update({hoho4: ppp4})
    #print "stop 5"


def daemon5(host, port, username, password, msisdn):
    #print "start 6"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    ppp5 = staros.outputparsing._getsubscribermsisdn(str(data))
    hoho5 = gethostname(host, port, username, password)
    if ppp5['sessnum'] != 0:
        out.update({hoho5: ppp5})
    #print "stop 6"


def daemon6(host, port, username, password, msisdn):
    #print "start 7"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    ppp6 = staros.outputparsing._getsubscribermsisdn(str(data))
    hoho6 = gethostname(host,port,username,password)
    if ppp6['sessnum']!=0:
        out.update({hoho6: ppp6})
    #print "stop 7"


def daemon7(host, port, username, password, msisdn):
    #print "start 8"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    ppp7 = staros.outputparsing._getsubscribermsisdn(str(data))
    hoho7 = gethostname(host, port, username, password)
    if ppp7['sessnum']!=0:
        out.update({hoho7: ppp7})
    #print "stop 8"


def daemon8(host, port, username, password, msisdn):
    #print "start 9"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    ppp8 = staros.outputparsing._getsubscribermsisdn(str(data))
    hoho8 = gethostname(host,port,username,password)
    if ppp8['sessnum']!=0:
        out.update({hoho8: ppp8})
    #print "stop 9"


def daemon9(host, port, username, password, msisdn):
    #print "start 10"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    ppp9 = staros.outputparsing._getsubscribermsisdn(str(data))
    hoho9 = gethostname(host,port,username,password)
    if ppp9['sessnum']!=0:
        out.update({hoho9: ppp9})
    #print "stop 10"


def daemon10(host, port, username, password, msisdn):
    #print "start 11"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    ppp10 = staros.outputparsing._getsubscribermsisdn(str(data))
    hoho10 = gethostname(host,port,username,password)
    if ppp10['sessnum']!=0:
        out.update({hoho10: ppp10})
    #print "stop 11"


def daemon11(host, port, username, password, msisdn):
    #print "start 12"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    ppp11 = staros.outputparsing._getsubscribermsisdn(str(data))
    hoho11 = gethostname(host,port,username,password)
    if ppp11['sessnum']!=0:
        out.update({hoho11: ppp11})
    #print "stop 12"


def daemon12(host,port,username,password,msisdn):
    #print "start 13"
    data = staros.connection._getsshdata3(host, port, username, password, "show subscribers msisdn " + msisdn + "\n")
    ppp12 = staros.outputparsing._getsubscribermsisdn(str(data))
    hoho12 = gethostname(host, port, username, password)
    if ppp12['sessnum']!=0:
        out.update({hoho12: ppp12})
    #print "stop 13"


def find(port, username, password, msisdn, ggsnlist):
    num = len(ggsnlist)

    if num == 1:
        a = threading.Thread(name='daemon', target=daemon(ggsnlist[0],port,username,password,msisdn))
        a.setDaemon(True)
        a.start()
        a.join()
    elif num == 2:
        a = threading.Thread(name='daemon', target=daemon, args=(ggsnlist[0],port,username,password,msisdn))
        b = threading.Thread(name='daemon1', target=daemon1, args=(ggsnlist[1],port,username,password,msisdn))
        a.setDaemon(True)
        b.setDaemon(True)
        a.start(),b.start()
        a.join(),b.join()
    elif num ==3:
        a = threading.Thread(name='daemon', target=daemon, args=(ggsnlist[0],port,username,password,msisdn))
        b = threading.Thread(name='daemon1', target=daemon1, args=(ggsnlist[1],port,username,password,msisdn))
        c = threading.Thread(name='daemon2', target=daemon2, args=(ggsnlist[2],port,username,password,msisdn))
        a.setDaemon(True),b.setDaemon(True),c.setDaemon(True)
        a.start(),b.start(),c.start()
        a.join(),b.join(),c.join()
    elif num ==4:
        a = threading.Thread(name='daemon', target=daemon, args=(ggsnlist[0],port,username,password,msisdn))
        b = threading.Thread(name='daemon1', target=daemon1, args=(ggsnlist[1],port,username,password,msisdn))
        c = threading.Thread(name='daemon2', target=daemon2, args=(ggsnlist[2],port,username,password,msisdn))
        d = threading.Thread(name='daemon3', target=daemon3, args=(ggsnlist[3],port,username,password,msisdn))
        a.setDaemon(True),b.setDaemon(True),c.setDaemon(True),d.setDaemon(True)
        a.start(),b.start(),c.start(),d.start()
        a.join(),b.join(),c.join(),d.join()
    elif num ==5:
        a = threading.Thread(name='daemon', target=daemon, args=(ggsnlist[0],port,username,password,msisdn))
        b = threading.Thread(name='daemon1', target=daemon1, args=(ggsnlist[1],port,username,password,msisdn))
        c = threading.Thread(name='daemon2', target=daemon2, args=(ggsnlist[2],port,username,password,msisdn))
        d = threading.Thread(name='daemon3', target=daemon3, args=(ggsnlist[3],port,username,password,msisdn))
        e = threading.Thread(name='daemon4', target=daemon4, args=(ggsnlist[4],port,username,password,msisdn))
        a.setDaemon(True),b.setDaemon(True),c.setDaemon(True),d.setDaemon(True),e.setDaemon(True)
        a.start(),b.start(),c.start(),d.start(),e.start()
        a.join(),b.join(),c.join(),d.join(),e.join()
    elif num ==6:
        a = threading.Thread(name='daemon', target=daemon, args=(ggsnlist[0],port,username,password,msisdn))
        b = threading.Thread(name='daemon1', target=daemon1, args=(ggsnlist[1],port,username,password,msisdn))
        c = threading.Thread(name='daemon2', target=daemon2, args=(ggsnlist[2],port,username,password,msisdn))
        d = threading.Thread(name='daemon3', target=daemon3, args=(ggsnlist[3],port,username,password,msisdn))
        e = threading.Thread(name='daemon4', target=daemon4, args=(ggsnlist[4],port,username,password,msisdn))
        f = threading.Thread(name='daemon5', target=daemon5, args=(ggsnlist[5],port,username,password,msisdn))
        a.setDaemon(True),b.setDaemon(True),c.setDaemon(True),d.setDaemon(True),e.setDaemon(True),f.setDaemon(True)
        a.start(),b.start(),c.start(),d.start(),e.start(),f.start()
        a.join(),b.join(),c.join(),d.join(),e.join(),f.join()
    elif num ==7:
        a = threading.Thread(name='daemon', target=daemon, args=(ggsnlist[0],port,username,password,msisdn))
        b = threading.Thread(name='daemon1', target=daemon1, args=(ggsnlist[1],port,username,password,msisdn))
        c = threading.Thread(name='daemon2', target=daemon2, args=(ggsnlist[2],port,username,password,msisdn))
        d = threading.Thread(name='daemon3', target=daemon3, args=(ggsnlist[3],port,username,password,msisdn))
        e = threading.Thread(name='daemon4', target=daemon4, args=(ggsnlist[4],port,username,password,msisdn))
        f = threading.Thread(name='daemon5', target=daemon5, args=(ggsnlist[5],port,username,password,msisdn))
        g = threading.Thread(name='daemon6', target=daemon6, args=(ggsnlist[6],port,username,password,msisdn))
        a.setDaemon(True),b.setDaemon(True),c.setDaemon(True),d.setDaemon(True),e.setDaemon(True),f.setDaemon(True),g.setDaemon(True)
        a.start(),b.start(),c.start(),d.start(),e.start(),f.start(),g.start()
        a.join(),b.join(),c.join(),d.join(),e.join(),f.join(),g.join()
    elif num ==8:
        a = threading.Thread(name='daemon', target=daemon, args=(ggsnlist[0],port,username,password,msisdn))
        b = threading.Thread(name='daemon1', target=daemon1, args=(ggsnlist[1],port,username,password,msisdn))
        c = threading.Thread(name='daemon2', target=daemon2, args=(ggsnlist[2],port,username,password,msisdn))
        d = threading.Thread(name='daemon3', target=daemon3, args=(ggsnlist[3],port,username,password,msisdn))
        e = threading.Thread(name='daemon4', target=daemon4, args=(ggsnlist[4],port,username,password,msisdn))
        f = threading.Thread(name='daemon5', target=daemon5, args=(ggsnlist[5],port,username,password,msisdn))
        g = threading.Thread(name='daemon6', target=daemon6, args=(ggsnlist[6],port,username,password,msisdn))
        h = threading.Thread(name='daemon7', target=daemon7, args=(ggsnlist[7],port,username,password,msisdn))
        a.setDaemon(True),b.setDaemon(True),c.setDaemon(True),d.setDaemon(True),e.setDaemon(True),f.setDaemon(True),g.setDaemon(True),h.setDaemon(True)
        a.start(),b.start(),c.start(),d.start(),e.start(),f.start(),g.start(),h.start()
        a.join(),b.join(),c.join(),d.join(),e.join(),f.join(),g.join(),h.join()
    elif num ==9:
        a = threading.Thread(name='daemon', target=daemon, args=(ggsnlist[0],port,username,password,msisdn))
        b = threading.Thread(name='daemon1', target=daemon1, args=(ggsnlist[1],port,username,password,msisdn))
        c = threading.Thread(name='daemon2', target=daemon2, args=(ggsnlist[2],port,username,password,msisdn))
        d = threading.Thread(name='daemon3', target=daemon3, args=(ggsnlist[3],port,username,password,msisdn))
        e = threading.Thread(name='daemon4', target=daemon4, args=(ggsnlist[4],port,username,password,msisdn))
        f = threading.Thread(name='daemon5', target=daemon5, args=(ggsnlist[5],port,username,password,msisdn))
        g = threading.Thread(name='daemon6', target=daemon6, args=(ggsnlist[6],port,username,password,msisdn))
        h = threading.Thread(name='daemon7', target=daemon7, args=(ggsnlist[7],port,username,password,msisdn))
        i = threading.Thread(name='daemon8', target=daemon8, args=(ggsnlist[8],port,username,password,msisdn))
        a.setDaemon(True),b.setDaemon(True),c.setDaemon(True),d.setDaemon(True),e.setDaemon(True),f.setDaemon(True),\
        g.setDaemon(True),h.setDaemon(True),i.setDaemon(True)
        a.start(),b.start(),c.start(),d.start(),e.start(),f.start(),g.start(),h.start(),i.start()
        a.join(),b.join(),c.join(),d.join(),e.join(),f.join(),g.join(),h.join(),i.join()
    elif num ==10:
        a = threading.Thread(name='daemon', target=daemon, args=(ggsnlist[0],port,username,password,msisdn))
        b = threading.Thread(name='daemon1', target=daemon1, args=(ggsnlist[1],port,username,password,msisdn))
        c = threading.Thread(name='daemon2', target=daemon2, args=(ggsnlist[2],port,username,password,msisdn))
        d = threading.Thread(name='daemon3', target=daemon3, args=(ggsnlist[3],port,username,password,msisdn))
        e = threading.Thread(name='daemon4', target=daemon4, args=(ggsnlist[4],port,username,password,msisdn))
        f = threading.Thread(name='daemon5', target=daemon5, args=(ggsnlist[5],port,username,password,msisdn))
        g = threading.Thread(name='daemon6', target=daemon6, args=(ggsnlist[6],port,username,password,msisdn))
        h = threading.Thread(name='daemon7', target=daemon7, args=(ggsnlist[7],port,username,password,msisdn))
        i = threading.Thread(name='daemon8', target=daemon8, args=(ggsnlist[8],port,username,password,msisdn))
        j = threading.Thread(name='daemon9', target=daemon9, args=(ggsnlist[9],port,username,password,msisdn))
        a.setDaemon(True),b.setDaemon(True),c.setDaemon(True),d.setDaemon(True),e.setDaemon(True),f.setDaemon(True),\
        g.setDaemon(True),h.setDaemon(True),i.setDaemon(True),j.setDaemon(True)
        a.start(),b.start(),c.start(),d.start(),e.start(),f.start(),g.start(),h.start(),i.start(),j.start()
        a.join(),b.join(),c.join(),d.join(),e.join(),f.join(),g.join(),h.join(),i.join(),j.join()
    elif num ==11:
        a = threading.Thread(name='daemon', target=daemon, args=(ggsnlist[0],port,username,password,msisdn))
        b = threading.Thread(name='daemon1', target=daemon1, args=(ggsnlist[1],port,username,password,msisdn))
        c = threading.Thread(name='daemon2', target=daemon2, args=(ggsnlist[2],port,username,password,msisdn))
        d = threading.Thread(name='daemon3', target=daemon3, args=(ggsnlist[3],port,username,password,msisdn))
        e = threading.Thread(name='daemon4', target=daemon4, args=(ggsnlist[4],port,username,password,msisdn))
        f = threading.Thread(name='daemon5', target=daemon5, args=(ggsnlist[5],port,username,password,msisdn))
        g = threading.Thread(name='daemon6', target=daemon6, args=(ggsnlist[6],port,username,password,msisdn))
        h = threading.Thread(name='daemon7', target=daemon7, args=(ggsnlist[7],port,username,password,msisdn))
        i = threading.Thread(name='daemon8', target=daemon8, args=(ggsnlist[8],port,username,password,msisdn))
        j = threading.Thread(name='daemon9', target=daemon9, args=(ggsnlist[9],port,username,password,msisdn))
        k = threading.Thread(name='daemon10', target=daemon10, args=(ggsnlist[10],port,username,password,msisdn))
        a.setDaemon(True),b.setDaemon(True),c.setDaemon(True),d.setDaemon(True),e.setDaemon(True),f.setDaemon(True),\
        g.setDaemon(True),h.setDaemon(True),i.setDaemon(True),j.setDaemon(True),k.setDaemon(True)
        a.start(),b.start(),c.start(),d.start(),e.start(),f.start(),g.start(),h.start(),i.start(),j.start(),k.start()
        a.join(),b.join(),c.join(),d.join(),e.join(),f.join(),g.join(),h.join(),i.join(),j.join(),k.join()
    elif num ==12:
        a = threading.Thread(name='daemon', target=daemon, args=(ggsnlist[0],port,username,password,msisdn))
        b = threading.Thread(name='daemon1', target=daemon1, args=(ggsnlist[1],port,username,password,msisdn))
        c = threading.Thread(name='daemon2', target=daemon2, args=(ggsnlist[2],port,username,password,msisdn))
        d = threading.Thread(name='daemon3', target=daemon3, args=(ggsnlist[3],port,username,password,msisdn))
        e = threading.Thread(name='daemon4', target=daemon4, args=(ggsnlist[4],port,username,password,msisdn))
        f = threading.Thread(name='daemon5', target=daemon5, args=(ggsnlist[5],port,username,password,msisdn))
        g = threading.Thread(name='daemon6', target=daemon6, args=(ggsnlist[6],port,username,password,msisdn))
        h = threading.Thread(name='daemon7', target=daemon7, args=(ggsnlist[7],port,username,password,msisdn))
        i = threading.Thread(name='daemon8', target=daemon8, args=(ggsnlist[8],port,username,password,msisdn))
        j = threading.Thread(name='daemon9', target=daemon9, args=(ggsnlist[9],port,username,password,msisdn))
        k = threading.Thread(name='daemon10', target=daemon10, args=(ggsnlist[10],port,username,password,msisdn))
        l = threading.Thread(name='daemon11', target=daemon11, args=(ggsnlist[11],port,username,password,msisdn))
        a.setDaemon(True),b.setDaemon(True),c.setDaemon(True),d.setDaemon(True),e.setDaemon(True),f.setDaemon(True),\
        g.setDaemon(True),h.setDaemon(True),i.setDaemon(True),j.setDaemon(True),k.setDaemon(True),l.setDaemon(True)
        a.start(),b.start(),c.start(),d.start(),e.start(),f.start(),g.start(),h.start(),i.start(),j.start(),k.start(),l.start()
        a.join(),b.join(),c.join(),d.join(),e.join(),f.join(),g.join(),h.join(),i.join(),j.join(),k.join(),l.join()
    return out


def find2(port, username, password, msisdn, ggsnlist):
    num = len(ggsnlist)

    threads = []

    for val in range(0, num):
        thread1 = threading.Thread(target=daemon, args=(ggsnlist[val], port, username, password, msisdn))
        #thread1.setDaemon(True)
        thread1.start()
        threads.append(thread1)

    for thread in threads:
        thread.join()

    return out


def find2imsi(port,username,password,imsi,ggsnlist):
    num = len(ggsnlist)

    threads = []

    for val in range(0, num):
        thread1 = threading.Thread(target=daemon1imsi, args=(ggsnlist[val], port, username, password, imsi))
        #thread1.setDaemon(True)
        thread1.start()
        threads.append(thread1)

    for thread in threads:
        thread.join()

    return out


def getenodeblist(host, port, username, password, excelpath, page, outputfile):
    s1u = ""
    book = openpyxl.load_workbook(excelpath)
    sheet = book[page]

    output = openpyxl.Workbook()
    sheet2 = output.active

    activestr = 2

    sheet2['A1'] = "peerid"
    sheet2['B1'] = "PLMN"
    sheet2['C1'] = "S1C-IP"
    sheet2['D1'] = "TAC"
    sheet2['E1'] = "BSC-number"
    sheet2['F1'] = "S1U-IP"


    data = connection._getsshdata3(host, port, username, password, "show mme-service enodeb-association all\n")
    #print data
    a1 = re.findall(r'\d\s+\d+\s+\S+\s+\S+\s+.\s\S+', data)

    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.load_system_host_keys()
    ssh.set_missing_host_key_policy(paramiko.WarningPolicy())
    conn = ssh.connect(host, port=port, username=username, password=password, timeout=5, look_for_keys=False)
    remote_conn = ssh.invoke_shell()






    for d in a1:
      try:
        sss = d.replace('  ',' ', 10)
        sss1 = sss.replace('  ', ' ', 10)
        sss2 = sss1.replace('  ', ' ', 10)
        ss = sss2.split(' ')
        peerid = ss[1]
        plmn = ss[2]
        bsc = plmn.split(":")[0][2:]
        s1u = ""

        if (bsc[0] == "0"):
            bsc = bsc[1:]

        try:

          for rowOfCellObjects in sheet['A1':'AK31']:
             for cellObj in rowOfCellObjects:
                dd = cellObj.value
                try:
                    if ("-" + bsc in dd):
                        #print(cellObj.coordinate, cellObj.value)

                        numlist = re.findall('\d+', cellObj.coordinate)
                        numlist2 = re.findall('\D+', cellObj.coordinate)
                        print  (numlist)
                        print (numlist2)

                        coord = numlist2[0] + str(int(numlist[0]) + 1)

                        #print(rowOfCellObjects[cellObj.coordinate + 1].value)
                        s1u = str(sheet[coord].value)


                except:
                    continue
        except:
            continue




        addr=ss[3].split(":")[0]
        #print plmn
        #print addr

        cnt = remote_conn.send("show mme-service enodeb-association full peer-id " + peerid + "\n")
        time.sleep(2)
        data7 = remote_conn.recv(10000)
        tac = re.findall(r'(\d{2,3}):\d{5}', data7)[0]



        print (peerid + " " + plmn + " " + addr + " " + tac + " " + bsc + " " + s1u)

        sheet2['A'+str(activestr)] = peerid
        sheet2['B' + str(activestr)] = plmn
        sheet2['C' + str(activestr)] = addr
        sheet2['D' + str(activestr)] = tac
        sheet2['E' + str(activestr)] = bsc
        sheet2['F' + str(activestr)] = s1u
        activestr = activestr+1

      except:
        continue


        #print ss[1]
    #print a1
    output.save(outputfile)


def getenodeblist2(host, port, username, password, excelpath, page, outputfile, lastsheet):

    book = openpyxl.load_workbook(excelpath)
    sheet = book[page]

    output = openpyxl.Workbook()
    sheet2 = output.active

    activestr = 2

    sheet2['A1'] = "peerid"
    sheet2['B1'] = "PLMN"
    sheet2['C1'] = "S1C-IP"
    sheet2['D1'] = "TAC"
    sheet2['E1'] = "BSC-number"
    sheet2['F1'] = "S1U-IP"
    sheet2['G1'] = "eNodeb-name"

    data = staros.connection._getsshdata4(host, port, username, password, "show mme-service enodeb-association full\n")
    splitdata = str(data).split("MMEMgr")
    print(splitdata)

    for i in splitdata:
        try:
            s1u = ""
            tac = re.findall(r'(\d{2,3}):25702', i)[0]
            enodebip = re.findall(r'eNodeB\s+IP\s+Address\s+:\s+(\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3})\s+', i)[0]
            enodebname = re.findall(r'eNodeB\s+Name\s+:\s+(\S+)', i)[0].replace('\\r\\n', '')
            peerid = re.findall(r'Peerid\s+:\s+(\S+)\s+', i)[0]
            plmn = re.findall(r'Global\s+ENodeB\s+ID\s+:\s+(\d+:\d+:\d+)', i)[0]
            bsc = plmn.split(":")[0][2:]
            if (bsc[0] == "0"):
                bsc = bsc[1:]
            try:
                for rowOfCellObjects in sheet['A1': lastsheet]:
                    for cellObj in rowOfCellObjects:
                        dd = cellObj.value
                        try:
                            if ("-"+bsc in dd):
                                # print(cellObj.coordinate, cellObj.value)
                                numlist = re.findall('\d+', cellObj.coordinate)
                                numlist2 = re.findall('\D+', cellObj.coordinate)
                                coord = numlist2[0] + str(int(numlist[0]) + 1)
                                # print(rowOfCellObjects[cellObj.coordinate + 1].value)
                                s1u = str(sheet[coord].value)


                        except:
                            continue
            except:
                continue

            print (peerid + " " + plmn + " " + enodebip + " " + tac + " " + enodebname + " " + bsc + " " + s1u)

            sheet2['A' + str(activestr)] = peerid
            sheet2['B' + str(activestr)] = plmn
            sheet2['C' + str(activestr)] = enodebip
            sheet2['D' + str(activestr)] = tac
            sheet2['E' + str(activestr)] = bsc
            sheet2['F' + str(activestr)] = s1u
            sheet2['G' + str(activestr)] = enodebname
            activestr = activestr + 1

        except:
            continue
    output.save(outputfile)
    #a1 = re.findall(r'\d\s+\d+\s+\S+\s+\S+\s+.\s\S+', data)


addrstatus = []

resultfinal = []

def getsgwadd(host, port, username, password, s1ulist):

    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.load_system_host_keys()
    ssh.set_missing_host_key_policy(paramiko.WarningPolicy())
    conn = ssh.connect(host, port=port, username=username, password=password, timeout=5, look_for_keys=False)
    remote_conn = ssh.invoke_shell()



    data = connection._getsshdata2(host, port, username, password, "context SPGW\nshow ip interface summary\n")
    s11ip = re.findall(r'S1[U,u]-loopback-\d{1,2}\s+(\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3})/32\s+Loopback\s+', data)
    #print s11ip
    for ipppppp in s11ip:
      for df in s1ulist:
        cmd = "\ncontext SPGW\nping " + df + " src " + ipppppp + " count 1\n"
        #print(cmd)
        cnt = remote_conn.send(cmd)
        time.sleep(5)
        output = remote_conn.recv(40000)
        time.sleep(1)

        res = re.findall(r'from\s(\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}):\sicmp.seq=\d+\sttl=\d+\stime=\d+.\d+\sms\sdscp=\d+', output)

        if (len(res) == 0):
            #print "failed"
            #print res
            #print output
            #print df
            addrstatus.append(df)
            resultfinal.append({df: {ipppppp: "failed"}})
            #print "ffff1"
            time.sleep(15)
        else:
            continue
            #resultfinal.append({df: "OK"})
        #print addrstatus
    #print s11ip



def ping_s1u_from_s11(host, port, username, password, sgwlistoam, excelpath, tac):
    excelpath = excelpath

    book = openpyxl.load_workbook(excelpath)

    sheet = book.active
    i = 2
    ress1u = []
    while i < 1000:
        ddd = sheet['F' + str(i)].value
        if (ddd != None):
            ress1u.append(ddd)
        i = i + 1
    book = openpyxl.load_workbook(excelpath)

    enodebaddrlist = []

    sheet = book.active
    i = 2
    while i < 1000:
        if (sheet['D' + str(i)].value == tac and sheet['F' + str(i)].value != None):
            enodebaddrlist.append(sheet['F' + str(i)].value)
        i = i + 1
    #print enodebaddrlist



    num = len(sgwlistoam)

    threads = []

    for val in range(0, num):
        thread1 = threading.Thread(target=getsgwadd, args=(sgwlistoam[val], port, username, password, enodebaddrlist))
        #thread1.setDaemon(True)
        thread1.start()
        threads.append(thread1)

    for thread in threads:
        thread.join()

    print (json.dumps(resultfinal, indent=4, sort_keys=True))

    return addrstatus

