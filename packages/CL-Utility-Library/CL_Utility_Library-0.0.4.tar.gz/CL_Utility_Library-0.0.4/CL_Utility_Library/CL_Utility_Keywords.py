import re
import xlrd
import openpyxl
import os
import shutil

def CheckEmail(email):
    if re.match("[\.\w]{2,}[@]\w+[.]\w+",email,re.IGNORECASE):
    	 return "PASS"
    else:
         return "FAIL"

def ContainsOnlyDigits(StrValue):
    # Using regex() 
    if re.match('^[0-9]*$', StrValue): 
       return   'PASS' 
    else: 
       return   'FAIL'

def ConvertStringToList(string): 
    ConvLst = list(string.split(" ")) 
    return ConvLst 

def GetValueInString(Value, Str): 
    Match = re.findall(Value, Str)
    if Match: 
       return    Match
    else: 
       return    'FAIL'
       
def FormatValidator(FieldFormat, FieldValue): 
    Regex_Match = re.match(FieldFormat, FieldValue)
    if Regex_Match: 
       return    'PASS'
    else: 
       return    'FAIL'

def GetStringPosition(str,fnd): 
    pos = 0
    ind = str.find(fnd)
    pos += str.find(fnd)
    return pos

def GetUniqueItems(iterable):
  for x in ints_list:
      if ints_list.count(x) > 1:
         ints_list.remove(x)
  return result

def DisposeWebDriver():
        """
        Closes the browser and shuts down the ChromeDriver executable
        that is started when starting the ChromeDriver
        """			
        try:
            self._debug('Closing all browsers')        
            browser.quit(self)
            browser.dispose(self)
            self._cache.close_all()
            self.empty_cache()
            RemoteWebDriver.close(self)
            Thread.Sleep(4000)
            RemoteWebDriver.quit(self)
            RemoteWebDriver.dispose(self)
            RemoteWebDrive = null
            WindowsUtils.KillProcessAndChildren("C:\Program Files\Python38\Scripts\chromedriver.exe") 
        except:
            # something probably has gone wrong
            pass
            
def RemoveSpecialCharacters(str):
     alphanumeric = ""
     for character in str:
        if character.isalnum():
           alphanumeric += character
     return	alphanumeric 


def CreateNewWorkbook(filename):
    #Create workbook object
    wb = openpyxl.Workbook()
    #Create Sheet and sheetname
    ws = wb.create_sheet("Sheet Name", 0)
    wb.save(filename)    
    
    
def OpenWorkbook(filename):
    #Open workbook
    wb = openpyxl.load_workbook(filename)
    for sheet in wb:
        print(sheet.title)

def GetTestCaseDataRow(testnme, excel_col, fileName, sheetname) :
    workbook = xlrd.open_workbook(fileName)
    worksheet = workbook.sheet_by_name(sheetname)
    rowEndIndex = worksheet.nrows - 1
    colEndIndex = worksheet.ncols - 1
    rowStartIndex = 1
    colStartIndex = 0
    testData = []
    dataRow = [] 
    cur_row = 0
    excel_col = int(excel_col)
    Found = 0
    for i in range(rowEndIndex):  
        cur_row = cur_row+1  
        testvalue = worksheet.cell_value(cur_row, excel_col)        
        if testvalue == testnme: 
            Found = 1
            break       
    if (Found == 0):
        return None  	    	                 
    cur_col = colStartIndex
    while cur_col <= colEndIndex:
         cell_type = worksheet.cell_type(cur_row, cur_col)
         value = worksheet.cell_value(cur_row, cur_col)
         dataRow.append(value)
         cur_col+=1              	     
    return dataRow
           

def GetDataRowCount(fileName, sheetname) :
    workbook = xlrd.open_workbook(fileName)
    worksheet = workbook.sheet_by_name(sheetname)
    rowEndIndex = worksheet.nrows - 1   
    return rowEndIndex

def GetDataByRowIndex(excel_row, filename, sheetname) :
    workbook = openpyxl.load_workbook(filename)
    worksheet =  workbook.get_sheet_by_name(sheetname)
    data_row = []                
    excel_row = int(excel_row)   
    for row in worksheet.iter_rows(min_row=excel_row, max_row=excel_row):
        for cell in row:
           #Append column values to the data row list
           data_row.append(cell.value)                                           	     
    return data_row  # return the row of test data
   
def GetNextAvailableDataRow(filename,sheetname):
     wb = openpyxl.load_workbook(filename)
     ws =  wb.get_sheet_by_name(sheetname)
     Available_Row = 0
     data_row = []
     i = 1    
     for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
         i = i + 1 
         if ws.cell(i, 26).value != "Used":
            Available_Row = i
            break   # exit for loop            
     for row in ws.iter_rows(min_row=Available_Row, max_row=Available_Row):
        for cell in row:
           data_row.append(cell.value)   #Append column values to the data row list                                             	     
     ws.cell(row=Available_Row, column=26).value = "Used" # Update 'Available Row' column cell value to 'Used'
     wb.save(filename) # Save the workbook 
     return data_row # return the row of data to the test case 

def GetAllDataFromExcelFile(fileName, sheetname) :
    workbook = xlrd.open_workbook(fileName)
    worksheet = workbook.sheet_by_name(sheetname)
    rowEndIndex = worksheet.nrows - 1
    colEndIndex = worksheet.ncols - 1
    rowStartIndex = 1
    colStartIndex = 0
    testData = []
    dataRow = []
    curr_row = rowStartIndex
    while curr_row <= rowEndIndex:
         cur_col = colStartIndex
         while cur_col <= colEndIndex:
             cell_type = worksheet.cell_type(curr_row, cur_col)

             value = worksheet.cell_value(curr_row, cur_col)
             dataRow.append(value)
             cur_col+=1
         curr_row += 1
         testData.append(dataRow)
    return dataRow
        
          