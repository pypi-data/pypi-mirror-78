from openpyxl import load_workbook
import os

class StdAcc:
    def __init__(self):
        self.std_file = "cpa_std/principal/account_std.xlsx"
        self.all=self._all()
    def _all(self):
        wb = load_workbook(self.std_file)
        l = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for r in range(2, ws.max_row + 1):
                _id=ws.cell(row=r, column=2).value
                _scope = ws.cell(row=r, column=4).value
                dic = dict(
                    id=_id,
                    name=ws.cell(row=r, column=3).value,
                    scope=_scope if _scope else "普遍适用",
                    kind=sheet,
                )
                if _id:
                    l.append(dic)
        return l
    @property
    def asset(self):
        return [i['name'] for i in self.all if i['kind']=="资产" or i['kind']=="资产或负债"]

    @property
    def liability(self):
        return [i['name'] for i in self.all if i['kind']=="负债" or i['kind']=="资产或负债"]
    @property
    def equity(self):
        return [i['name'] for i in self.all if i['kind']=="权益"]

    @property
    def income(self):
        return [i['name'] for i in self.all if i['kind'] == "利润表"]
    @property
    def cashFlow(self):
        return [i['name'] for i in self.all if i['kind'] == "现金流量表"]
if __name__ == '__main__':
    s=StdAcc()
    s.all
