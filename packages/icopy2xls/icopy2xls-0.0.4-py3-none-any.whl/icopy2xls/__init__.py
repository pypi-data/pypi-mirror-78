import sys
import json
import xlrd3 as xlrd
from openpyxl import load_workbook
from logging import getLogger

logger = getLogger("pyC")


class Writer:
    """
    This class is made to update existing excel file.
    First it will open the file in python and then we can do multiple writes and once everything is update we can use
    save method in order to save the updated excel file. Hence, this class is very useful is saving time while updating
    excel files.
    """
    def __init__(self, path, sheet):
        self.path = path
        self.sht = sheet
        self.workbook = load_workbook(self.path)
        self.sheet = self.workbook[self.sht]

    def write(self, row, col, data):
        # Update the values using row and col number.
        # Note :- We are using openpyxl so row & column index will start from 1 instead of 0
        self.sheet.cell(row, col).value = data

    def save(self):
        # Call this method to save the file once every updates are written
        self.workbook.save(self.path)
        self.workbook.close()


class Mover:
    """
    This class is used to copy data from one excel file to another. We can also use some filters to copy only targeted
    data.
    """
    def __init__(self, source_file_path, source_sheet, destination_file_path, destination_sheet, lines=0):
        """

        :param source_file_path: Path with filename of source excel file
        :param source_sheet: Sheet name of the source
        :param destination_file_path: Path with filename of Destination excel file
        :param destination_sheet: Sheet name of the destination
        :param lines: Number of lines starting from 1 to be considered for moving from source.
        """
        self.source_file_path = source_file_path
        if not isinstance(self.source_file_path, list):
            self.source_file_path = [self.source_file_path]
        self.source_sheet = source_sheet
        self.destination_file_path = destination_file_path
        self.destination_sheet = destination_sheet
        self.lines = self.process_lines(lines)

    def move(self, filters={}, add_missing_columns=False):
        """

        :param filters: dictionary of filters with header as key and data as value or list of data as value
        :param add_missing_columns: True if we need all headers from source file to destination file which are currently
        not in destination file.
        :return:
        """
        for source_file_path in self.source_file_path:
            try:
                source = self.read_xlsx(source_file_path, self.source_sheet)
                destination = self.read_xlsx(self.destination_file_path, self.destination_sheet)
            except FileNotFoundError as e:
                logger.critical(f"File not found: {e}")
                return
            destination_wb = Writer(self.destination_file_path, self.destination_sheet)  # Writer class object used to update existing file
            if add_missing_columns:
                self.add_missing_columns(source, destination, destination_wb)
                # again opening destination file as it is updated with the source file columns
                destination = self.read_xlsx(self.destination_file_path, self.destination_sheet)

            end = self.lines  # if number of rows to be considered from source is not predefined the take all
            if not end:
                for i in range(1, source.nrows):
                    end.append(i)
            new_data = [source.row(row) for row in end]  # create a new list of all data and remove the filtered data
            remove_data = []  # rows not matching filter are stored here which is used later to remove data from new_data
            for filter in filters:  # iterate through the dictionary of filter and
                ind = [x.value for x in source.row(0)].index(filter)  # getting index of filter header then use the same index to check data
                if type(filters[filter]) is not list:
                    filters[filter] = [filters[filter]]
                for row in new_data:
                    if row[ind].value not in filters[filter]:  # check if data is matching with filter
                        remove_data.append(row)
                for row in remove_data:  # removing unmatched data from new_data list
                    try:
                        new_data.remove(row)
                    except ValueError:
                        pass
            row_num = destination.nrows  # used to maintain new row number
            for data in new_data:  # iterating through the data to be written and writing then on the correct cells
                row_num += 1
                for cell in range(len(data)):
                    try:
                        # getting column number where new data is to be written with the help of indexing header in destination file
                        ind = [x.value for x in destination.row(0)].index(source.row(0)[cell].value) + 1
                    except ValueError:
                        # if add_missing_columns is false then ValueError is thrown for the headers which are not present in destination
                        continue
                    destination_wb.write(row_num, ind, data[cell].value)
            destination_wb.save()

    def read_xlsx(self, path, sheet):
        # reading xlsx file using xlrd
        wb = xlrd.open_workbook(path)
        sht = wb.sheet_by_name(sheet)
        return sht

    def add_missing_columns(self, source, destination, destination_wb):
        # looking for the headers which are not present in destination file and then updating destination file
        source_headers = [x.value for x in source.row(0)]
        destination_headers = [x.value for x in destination.row(0)]
        col = len(destination_headers)
        for headers in source_headers:
            if headers not in destination_headers:
                col += 1
                destination_wb.write(1, col, headers)
        destination_wb.save()

    def process_lines(self, lines):
        line_lis = []
        if isinstance(lines, int):
            if not lines < 1:
                for i in range(1, lines+1):
                    line_lis.append(i)
            return line_lis
        lis = [x.strip() for x in lines.strip().split(',')]
        while "" in lis:
            lis.remove("")
        for l in lis:
            if "-" in l:
                start = int(l.split("-")[0].strip())
                end = int(l.split("-")[1].strip()) + 1
                for i in range(int(start), int(end)):
                    if i not in line_lis:
                        line_lis.append(i)
            elif l.strip().isnumeric():
                i = int(l.strip())
                if i not in line_lis:
                    line_lis.append(i)
            else:
                print(l)
                raise BaseException('Lines structure is wrong. For multiple values "," is used & for range "-" is used'\
                                    'please verify it again!')
        return line_lis


def parse_json(path):
    """
    Takes parameter from a json file.
    :param path:
    :return:
    """
    js = json.load(open(path))
    lines = int(js["lines"]) or 0
    mover = Mover(js["source_file"], js["source_sheet"], js["destination_file"], js["destination_sheet"], lines)
    if js["add_missing_columns"].lower() == "true":
        add_missing = True
    else:
        add_missing = False
    mover.move(js["filters"], add_missing)


if __name__ == "__main__":
    js = sys.argv[-1]
    if ".json" in js.lower():
        parse_json(js)
