"""
Extensions to `openpyxl` classes allowing for the cells corresponding to table columns
to be returned.
"""

from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Union

from openpyxl.cell.cell import Cell as openpyxl_Cell
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.workbook import Workbook as openpyxl_Workbook
from openpyxl.worksheet.table import Table as openpyxl_Table
from openpyxl.worksheet.table import TableColumn as openpyxl_TableColumn
from openpyxl.worksheet.worksheet import Worksheet as openpyxl_Worksheet


class Meta:
    """
    Iterate over children.
    """

    def __init__(self, children: Optional[Union[Dict[str, Any], List[openpyxl_Cell]]]):
        self._children = children

    def __iter__(self):
        return iter(self._children.values())

    def __getitem__(self, index):
        return self._children[index]


class Book(Meta):
    """
    Extends the `openpyxl.workbook.workbook.Workbook` class from `openpyxl`.
    """

    def __init__(self, openpyxl_book: openpyxl_Workbook):
        self.openpyxl_book = openpyxl_book
        self.sheets = {
            openpyxl_sheet.title: Sheet(openpyxl_sheet, parent=self)
            for openpyxl_sheet in openpyxl_book.worksheets
        }
        super().__init__(children=self.sheets)


class Sheet(Meta):
    """
    Extends the `openpyxl.worksheet.worksheet.Worksheet` class from `openpyxl`.
    """

    def __init__(self, openpyxl_sheet: openpyxl_Worksheet, parent: Book):
        self.parent = parent
        self.openpyxl_sheet = openpyxl_sheet
        self.tables = {
            openpyxl_table.name: Table(openpyxl_table, parent=self)
            for openpyxl_table in openpyxl_sheet.tables.values()
        }
        super().__init__(children=self.tables)


class Table(Meta):
    """
    Extends the `openpyxl.worksheet.table.Table` class from `openpyxl`.
    """

    def __init__(self, openpyxl_table: openpyxl_Table, parent: Sheet):
        self.parent = parent
        self.openpyxl_table = openpyxl_table
        (self.first_col, self.first_row, _, self.last_row) = range_boundaries(
            openpyxl_table.ref
        )
        self.columns: Dict[str, Column] = {}
        super().__init__(children=self.columns)

        for openpyxl_column in openpyxl_table.tableColumns:
            self.columns[openpyxl_column.name] = Column(self, openpyxl_column)

        self.cells: Optional[List[ColumnCells]] = None

    def get_cells(self):
        """
        Get the cells in the table as a list of `ColumnCells` objects.
        """
        table_cells = [column.get_cells() for column in self]
        self.cells = {
            column.openpyxl_column.name: column_cells
            for (column, column_cells) in zip(self, table_cells)
        }
        return self.cells


class Column(Meta):
    """
    Extends the `openpyxl.worksheet.table.TableColumn` class from `openpyxl`.
    """

    def __init__(self, parent: Table, openpyxl_column: openpyxl_TableColumn):
        self.parent = parent
        self.openpyxl_column = openpyxl_column

        self.col_num = parent.first_col + len(parent.columns)

        self.cells: Optional[ColumnCells] = None
        super().__init__(children=None)

    def get_cells(self):
        """
        Get cells in this column as a `ColumnCells` object.
        """

        table = self.parent
        header_row_count = table.openpyxl_table.headerRowCount
        totals_row_count = table.openpyxl_table.totalsRowCount

        sheet = table.parent

        if table.openpyxl_table.headerRowCount == 1:
            header = sheet.openpyxl_sheet.cell(row=table.first_row, column=self.col_num)
            header_row_count = 1
        else:
            header = None
            header_row_count = 0

        if table.openpyxl_table.totalsRowCount == 1:
            total = sheet.openpyxl_sheet.cell(row=table.last_row, column=self.col_num)
            totals_row_count = 1
        else:
            total = None
            totals_row_count = 0

        between = next(
            sheet.openpyxl_sheet.iter_cols(
                min_col=self.col_num,
                max_col=self.col_num,
                min_row=table.first_row + header_row_count,
                max_row=table.last_row - totals_row_count,
            )
        )

        self.cells = ColumnCells(header, between, total)
        return self.cells


@dataclass
class ColumnCells:
    """
    Contains a column of cells from an Excel table.
    """

    header: Optional[openpyxl_Cell]
    between: List[openpyxl_Cell]
    total: Optional[openpyxl_Cell]
