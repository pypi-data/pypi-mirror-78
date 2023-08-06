import logging
from pathlib import Path
from time import time
import os, re, json, pickle, base64, queue
from threading import Thread
from sys import exit  as s_exit


def divide_into_equal_parts(sequence, parts_n) -> list:
    n_averge = len(sequence) // parts_n
    divided_sequence = []
    for x in range(parts_n - 1):
        divided_sequence.append(sequence[x * n_averge:(x + 1) * n_averge])

    divided_sequence.append(sequence[n_averge * (parts_n - 1):])
    return divided_sequence


def NewThread(func_names=None, func_params=None, dont_join=True):
    func_names = func_names or []
    func_params = func_params or []
    if callable(func_names):
        func_names = [func_names for _ in range(len(func_params))]
    func_n = len(func_names)
    if not func_n:
        print('Please enter functions into the func_names,the system will exit!')
        s_exit()
    threadings_n = len(func_params)
    if func_n != 1 and threadings_n != func_n:
        print('The params\' count is not equal the functions name\'count,the system will exit!')
        s_exit()
    if func_n == 1:
        func_names = [func_names[0] for _ in range(threadings_n)]
    threads = []
    for f, p in zip(func_names, func_params):
        p = p if type(p) == tuple else tuple([p])
        if p == (None,):
            p = ()
        a = Thread(target=f, args=p)
        a.start()
        threads.append(a)
    if not dont_join:
        for x in threads:
            x.join()
    else:
        return threads


# File Options
class File:

    @staticmethod
    def save_to_file(obj, filename):
        method = 'wb' if type(obj) == bytes else 'w'
        with open(filename, method)as f:
            f.write(obj)

    @staticmethod
    def load_from_file(filename):
        try:
            with open(filename, 'r') as f:
                content = f.read()
        except UnicodeDecodeError:
            with open(filename, 'rb') as f:
                content = f.read()
        return content

    @staticmethod
    def save_to_json(obj, filename):
        with open(filename, 'w')as f:
            json.dump(obj, f)

    @staticmethod
    def save_to_pickle(obj, filename):
        with open(filename, 'wb')as f:
            pickle.dump(obj, f)

    @staticmethod
    def load_from_json(filename):
        with open(filename, 'r')as f:
            return json.load(f)

    @staticmethod
    def load_from_pickle(filename):
        with open(filename, 'rb')as f:
            return pickle.load(f)

    @staticmethod
    def save_to_excel(data, fn='data.xlsx', first_col_name=None, auto_adjust_col_width=True):
        import pandas

        # return if data is empty
        if not data:
            return

        result = pandas.DataFrame(columns=([first_col_name] + list(data[0].keys()))) if not os.path.exists(
            fn) else pandas.read_excel(fn)
        for i, d in enumerate(data, start=len(result[result.columns[0]])):
            dict_d = {}
            if first_col_name is not None:
                dict_d.update({first_col_name: i})

            if d:
                dict_d.update({k: [v] for k, v in d.items()})
            else:
                continue
            result = result.append(pandas.DataFrame(dict_d, index=[0], ), sort=False)

        result = result.loc[:, [first_col_name] + list(data[0].keys())]

        result = result.dropna(1, 'all')
        result.to_excel(fn, index=False)

        if auto_adjust_col_width:
            File.adjust_excel_col_width(fn)

    @staticmethod
    def adjust_excel_col_width(fn):
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        wb = load_workbook(fn)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            col_max_width = [max(
                [len(str(r.value)) + sum(len(x) for x in re.findall('[^\x00-\xff]+', str(r.value))) if r.value else 0
                 for r in col]) for col in sheet.columns]

            for i, w in enumerate(col_max_width):
                col_letter = get_column_letter(i + 1)
                sheet.column_dimensions[col_letter].width = max(min(w * 2, 100), 10)

        wb.save(filename=fn.replace('.', '_adj.'))


# Encryption Options
class Encrypt:

    @staticmethod
    def b64_encode(str_in):
        return base64.b64encode(str_in.encode('utf-8')).decode('utf-8')

    @staticmethod
    def b64_decode(str_in):
        return base64.b64decode(str_in).decode('utf-8')

    @staticmethod
    def md5(cnt):
        import hashlib
        hl = hashlib.md5()
        hl.update(str(cnt).encode('utf-8'))
        return hl.hexdigest()


# Time Options
def time_counter(n):
    now = time()

    def creat():
        nonlocal now
        if time() - now > n:
            now = time()
            return True
        return False

    return creat


class timer:
    n = 0

    def __enter__(self):
        self.time_begin = time()

    def __exit__(self, type, value, traceback):
        timer.n = time() - self.time_begin


class Timer:
    def __enter__(self):
        self.time_start = time()
        return self

    def __exit__(self, exe_type, exe_value, exe_traceback):
        pass

    def get_time_consuming(self, flush_time_start=False):
        now = time()
        ts = now - self.time_start
        if flush_time_start:
            self.time_start = now
        return ts


# List,str,dict Options
def unique_list(l, delete=True):
    u_l = []
    for x in l:
        if x not in u_l: u_l.append(x)
    if delete: del (l)
    return u_l


class SafeQueue(queue.Queue):
    def get(self, key, none_if_get_unexpected_key=False):
        while True:
            value = super(SafeQueue, self).get()
            try:
                k, v = value
            except Exception as e:
                print(e)
                continue
            if k == key or key == '*':
                return v
            elif none_if_get_unexpected_key:
                print('Unexpected key ,value-> %s:%s' % (k, v))
                return None
            else:
                print('Unexpected key ,value-> %s:%s' % (k, v))

    def clear(self):
        while not self.empty():
            self.get('*')

    def safe_get(self, key):
        self.clear()
        return self.get(key)


class DictQueue:
    def __init__(self, init_keys: list, get_callback=None, put_callback=None):
        self.dict = {k: queue.Queue() for k in init_keys}
        self.get_callback = get_callback
        self.put_callback = put_callback

    def put(self, k_v: list):
        k, v = k_v
        self.dict[k].put(v)
        if self.put_callback is not None:
            self.put_callback(k, v)

    def get(self, key):
        value = self.dict[key].get()
        if self.get_callback is not None:
            self.get_callback(key, value)

        return value


def string_similar_rate(s1, s2):
    import difflib
    return difflib.SequenceMatcher(None, s1, s2).quick_ratio()


class IgnoreFilter(logging.Filter):
    def __init__(self, keyword):
        logging.Filter.__init__(self)
        self.keyword = keyword

    def filter(self, record):
        if record.stack_info:
            record.stack_info = re.sub('File.+idlelib/run.py.+?\n.+?\n', '',
                                       record.stack_info, flags=re.S)
            record.stack_info = re.sub(' +(?=File)', ' ' * 2, record.stack_info)
            record.stack_info = re.sub(' {2}File.+?\n.+?self\.logger\.warning\(msg,stack_info=stack_info.+?$',
                                       '', record.stack_info)
        return True


class Log:
    NOSET, DEBUG, INFO, WARN, ERROR = logging.NOTSET, logging.DEBUG, logging.INFO, logging.WARN, logging.ERROR
    separator = ' '.join([' ', '-' * 3, ' '])

    def __init__(self, name, ch_level=logging.DEBUG, fh_level=None, log_filepath=None):
        logger = logging.getLogger(name)
        logger.propagate = False
        logger.setLevel(logging.DEBUG)
        ch = logging.StreamHandler()
        ch.setLevel(ch_level)
        formatter = logging.Formatter('\n>>> %(levelname)s:%(name)s:%(asctime)s:\n    %(message)s', "%Y/%m/%d %H:%M:%S")
        if fh_level is not None:
            if log_filepath is None:
                log_filepath = Path('Log', name + '.log')
                if not log_filepath.parent.is_dir():
                    log_filepath.parent.mkdir()

            log_filepath.touch()
            fh = logging.FileHandler(log_filepath)
            fh.setLevel(fh_level)
            fh.setFormatter(formatter)
            logger.addHandler(fh)

        ch.setFormatter(formatter)
        # ch.addFilter(IgnoreFilter(''))
        logger.addHandler(ch)
        self.logger = logger

    def setLevel(self, level):
        self.logger.setLevel(level)

    def debug(self, *args, **kw):
        msg = self.separator.join([str(x) for x in args])
        self.logger.debug(msg, **kw)

    def info(self, *args, **kw, ):
        msg = self.separator.join([str(x) for x in args])
        self.logger.info(msg, **kw)

    def warning(self, *args, stack_info=True, **kw, ):
        msg = self.separator.join([str(x) for x in args])
        self.logger.warning(msg, stack_info=stack_info, **kw)

    def error(self, *args, stack_info=True, **kw, ):
        msg = self.separator.join([str(x) for x in args])
        self.logger.error(msg, stack_info=stack_info, **kw)
